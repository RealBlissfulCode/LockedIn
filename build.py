import json
# -*- coding: utf-8 -*-
import csv, json, html, os
from ingredients import ING
from recipes_1 import RECIPES_1
from recipes_2 import RECIPES_2
from recipes_3 import RECIPES_3
from recipes_4 import RECIPES_4
from recipes_5 import RECIPES_5
from recipes_6 import RECIPES_6
from theme import CSS as THEME_CSS, EXTRA as THEME_EXTRA, EXTRA2 as THEME_EXTRA2, EXTRA3 as THEME_EXTRA3, EXTRA4 as THEME_EXTRA4, EXTRA5 as THEME_EXTRA5, EXTRA6 as THEME_EXTRA6
import formulas as FRM
import calculator as CALCU
import ingredient_list as ILIST
import collections_page as COLL
import voice as V
import prices as PR
import app as APP
import household as HH
import content as C

RECIPES = RECIPES_1 + RECIPES_2 + RECIPES_3 + RECIPES_4 + RECIPES_5 + RECIPES_6
OUT = "/mnt/user-data/outputs"
os.makedirs(OUT, exist_ok=True)

NUTS = ["kcal","p","c","f","fib","leu","ca","fe","mg","k","zn","na","b12","vitd","ala","epadha"]
DV = {"ca":1300,"fe":18,"mg":420,"k":4700,"zn":11,"b12":2.4,"vitd":20,"na":2300}
DVNAME = {"ca":"Calcium","fe":"Iron","mg":"Magnesium","k":"Potassium","zn":"Zinc",
          "b12":"Vitamin B12","vitd":"Vitamin D"}

def compute(r):
    tot = {n:0.0 for n in NUTS}; grams = 0.0
    for key, g, _m in r["ing"]:
        if key not in ING: raise KeyError(f"{r['id']}: unknown ingredient {key}")
        item = ING[key]; grams += g
        for n in NUTS: tot[n] += item[n] * g / 100.0
    s = r["servings"]
    per = {n: tot[n]/s for n in NUTS}
    per["grams"] = grams/s
    return per

def satiety(per, cat):
    sc = 0.30*per["p"] + 0.60*per["fib"] + 0.01*per["grams"] - (6 if cat=="Drink" else 0)
    lvl = "HIGH" if sc>=16 else ("MEDIUM" if sc>=9 else "LOWER")
    return round(sc,1), lvl

def difficulty(r):
    t = r["prep"] + r["cook"]
    return ("EASY" if t<=10 else ("MODERATE" if t<=25 else "ADVANCED")), t

def autotags(r, per, sat_lvl):
    t=[]
    kcal=per["kcal"]; p=per["p"]; c=per["c"]; f=per["f"]; fib=per["fib"]; leu=per["leu"]
    if p>=40: t.append("VERY HIGH PROTEIN")
    elif p>=30: t.append("HIGH PROTEIN")
    if leu>=3.0: t.append("LEUCINE PRIORITY")
    elif leu>=2.5: t.append("HIGH LEUCINE")
    if kcal>=500: t.append("HIGH CALORIE")
    if kcal<=300: t.append("LOW CALORIE")
    if fib>=8: t.append("HIGH FIBER")
    if f>=20: t.append("HIGH HEALTHY FAT")
    if c>=50 and kcal>0 and (c*4)/kcal>=0.50: t.append("HIGH CARB")
    if c<=25: t.append("LOW CARB")
    if kcal>0:
        pp,cp,fp = p*4/kcal, c*4/kcal, f*9/kcal
        if 0.20<=pp<=0.40 and 0.30<=cp<=0.55 and 0.20<=fp<=0.40: t.append("BALANCED MACRO")
    if p>=30 and leu>=2.5: t.append("MUSCLE-BUILDING PRIORITY")
    if p>=25 and c>=40: t.append("POST-WORKOUT FRIENDLY")
    if c>=30 and f<15 and fib<6: t.append("PRE-WORKOUT FRIENDLY")
    if sat_lvl=="HIGH": t.append("HIGH SATIETY")
    d,_ = difficulty(r)
    if d=="EASY" and r["prep"]+r["cook"]<=10: t.append("QUICK")
    if kcal>=600 and p>=35: t.append("TWO-MEAL-DAY FRIENDLY")
    for k in ["ca","fe","mg","k","zn"]:
        if per[k]/DV[k] >= 0.20: t.append("HIGH "+{"ca":"CALCIUM","fe":"IRON","mg":"MAGNESIUM","k":"POTASSIUM","zn":"ZINC"}[k])
    if per["epadha"]>=0.5 or per["ala"]>=2.5: t.append("OMEGA-3 RICH")
    hits = sum(1 for k in ["ca","fe","mg","k","zn","b12","vitd"] if per[k]/DV[k]>=0.20)
    if hits>=4: t.append("HIGH MICRONUTRIENT DENSITY")
    if r["cat"]!="SDA Meat/Fish": t.append("VEGETARIAN")
    t.append("SDA-COMPATIBLE"); t.append("GLUTEN-FREE")
    for m in r.get("manual_tags",[]):
        if m not in t: t.append(m)
    seen=set(); out=[]
    for x in t:
        if x not in seen: seen.add(x); out.append(x)
    return out

def micros(per):
    out=[]
    for k,label in DVNAME.items():
        pct = per[k]/DV[k]*100
        if pct>=15: out.append(f"{label} {int(round(pct))}% DV")
    if per["epadha"]>=0.2: out.append(f"EPA+DHA {per['epadha']:.1f} g")
    if per["ala"]>=1.0: out.append(f"ALA omega-3 {per['ala']:.1f} g")
    if per["na"]>=800: out.append(f"Sodium {int(per['na'])} mg (high)")
    return out

def validate(r, per):
    issues=[]
    at = per["p"]*4 + per["c"]*4 + per["f"]*9 - per["fib"]*2
    if per["kcal"]>0:
        dev = abs(per["kcal"]-at)/per["kcal"]
        if dev>0.15: issues.append(f"{r['id']} Atwater deviation {dev*100:.0f}% (kcal {per['kcal']:.0f} vs {at:.0f})")
    if per["p"]>0 and per["leu"]/per["p"] > 0.13: issues.append(f"{r['id']} leucine fraction implausible")
    if not r["steps"]: issues.append(f"{r['id']} missing steps")
    if not r["ing"]: issues.append(f"{r['id']} missing ingredients")
    used = {k for k,_,_ in r["ing"]}
    return issues

DATA=[]; ISSUES=[]
for r in RECIPES:
    per = compute(r)
    sc,lvl = satiety(per, r["cat"])
    d,tt = difficulty(r)
    tags = autotags(r, per, lvl)
    ISSUES += validate(r, per)
    DATA.append(dict(r=r, per=per, sat=sc, satlvl=lvl, diff=d, time=tt, tags=tags, micros=micros(per)))

ids=[d["r"]["id"] for d in DATA]
assert len(ids)==len(set(ids)), "duplicate IDs"
print("Recipes:", len(DATA))
from collections import Counter
print(Counter(d["r"]["cat"] for d in DATA))
print("VALIDATION ISSUES:", len(ISSUES))
for i in ISSUES: print("  -", i)

# ---------------- CSV / XLSX ----------------
HEAD = ["ID","Name","Category","Difficulty","TotalMin","Servings","Calories","Protein_g","Carbs_g",
        "Fat_g","Fiber_g","Leucine_g","SatietyScore","Satiety",
        "CostWalmart_total","CostWalmart_serving","CostCostco_total","CostCostco_serving",
        "TotalKcal_allServings","TotalProtein_allServings","Tags"]
rows=[]
for d in DATA:
    p=d["per"]
    rows.append([d["r"]["id"], d["r"]["name"], d["r"]["cat"], d["diff"], d["time"], d["r"]["servings"],
                 int(round(p["kcal"]/5.0)*5), round(p["p"]), round(p["c"]), round(p["f"]),
                 round(p["fib"]), round(p["leu"],1), d["sat"], d["satlvl"],
                 round(PR.cost_of(d["r"], d["r"]["servings"], "walmart")[0],2),
                 round(PR.cost_of(d["r"], d["r"]["servings"], "walmart")[1],2),
                 round(PR.cost_of(d["r"], d["r"]["servings"], "costco")[0],2),
                 round(PR.cost_of(d["r"], d["r"]["servings"], "costco")[1],2),
                 int(round(p["kcal"]*d["r"]["servings"])), round(p["p"]*d["r"]["servings"]),
                 "; ".join(d["tags"])])
with open(f"{OUT}/meal_database.csv","w",newline="") as fh:
    w=csv.writer(fh); w.writerow(HEAD); w.writerows(rows)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
wb=Workbook(); ws=wb.active; ws.title="Meal Database"
ws.append(HEAD)
for c in ws[1]:
    c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="12365F")
    c.alignment=Alignment(vertical="center",wrap_text=True)
for row in rows: ws.append(row)
for r_ in ws.iter_rows(min_row=2):
    for c in r_: c.font=Font(name="Arial")
widths=[8,44,15,12,10,10,10,11,10,9,10,11,13,10,15,16,15,16,17,18,70]
for i,w_ in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w_
ws.freeze_panes="C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{len(rows)+1}"

ws2=wb.create_sheet("Targets and Formulas")
ws2["A1"]="JARON - NUTRITION TARGETS"; ws2["A1"].font=Font(name="Arial",bold=True,size=14)
inputs=[("Bodyweight (lb)",147.3),("Height (in)",67),("Age",20),("Activity factor",1.725),
        ("Protein factor (g per lb)",1.05),("Fat factor (g per lb)",0.6),("Surplus (kcal)",300)]
ws2["A3"]="INPUTS - edit the blue cells"; ws2["A3"].font=Font(name="Arial",bold=True)
r0=4
for i,(lab,val) in enumerate(inputs):
    ws2.cell(r0+i,1,lab).font=Font(name="Arial")
    cc=ws2.cell(r0+i,2,val); cc.font=Font(name="Arial",color="0000FF"); cc.fill=PatternFill("solid",fgColor="FFFF00")
calc_start=r0+len(inputs)+2
ws2.cell(calc_start-1,1,"CALCULATED").font=Font(name="Arial",bold=True)
calcs=[("Bodyweight (kg)","=B4/2.2046"),("Height (cm)","=B5*2.54"),
       ("BMR (Mifflin-St Jeor)","=10*B{kg}+6.25*B{cm}-5*B6+5".format(kg=calc_start,cm=calc_start+1)),
       ("Maintenance (TDEE)","=B{}*B7".format(calc_start+2)),
       ("Goal calories","=B{}+B10".format(calc_start+3)),
       ("Protein (g)","=B4*B8"),("Fat (g)","=B4*B9"),
       ("Carbs (g)","=(B{gc}-B{pg}*4-B{fg}*9)/4".format(gc=calc_start+4,pg=calc_start+5,fg=calc_start+6)),
       ("Fiber (g)","=B{}/1000*14".format(calc_start+4)),
       ("Water (ml)","=B{}*35+750".format(calc_start))]
for i,(lab,frm) in enumerate(calcs):
    ws2.cell(calc_start+i,1,lab).font=Font(name="Arial")
    ws2.cell(calc_start+i,2,frm).font=Font(name="Arial")
ws2.cell(calc_start+len(calcs)+1,1,"Source: Mifflin-St Jeor equation. Bodyweight from Hume Pod, May 3 2026. Activity factor 1.725 assumes a physical job plus daily training.").font=Font(name="Arial",italic=True,size=9)
ws2.column_dimensions["A"].width=32; ws2.column_dimensions["B"].width=16

ws3=wb.create_sheet("Ingredient Data per 100g")
ws3.append(["Ingredient key","Name","kcal","Protein g","Carb g","Fat g","Fiber g","Leucine g",
            "Calcium mg","Iron mg","Magnesium mg","Potassium mg","Zinc mg","Sodium mg","B12 mcg","VitD mcg","ALA g","EPA+DHA g"])
for c in ws3[1]:
    c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="12365F")
for k,v in sorted(ING.items()):
    ws3.append([k,v["name"]]+[v[n] for n in NUTS])
for r_ in ws3.iter_rows(min_row=2):
    for c in r_: c.font=Font(name="Arial")
ws3.column_dimensions["A"].width=24; ws3.column_dimensions["B"].width=46
_ihtml,_irows,_ihead = ILIST.build(ING, DATA, OUT)
ws4=wb.create_sheet("Shopping and Ingredients")
ws4.append(_ihead)
for c in ws4[1]:
    c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="12365F")
    c.alignment=Alignment(vertical="center",wrap_text=True)
for row in _irows: ws4.append(row)
for r_ in ws4.iter_rows(min_row=2):
    for c in r_: c.font=Font(name="Arial")
for i,w_ in enumerate([34,34,24,13,15,15,12,12,12,10,10,12,13,13,60,50],1):
    ws4.column_dimensions[get_column_letter(i)].width=w_
ws4.freeze_panes="C2"
ws4.auto_filter.ref=f"A1:{get_column_letter(len(_ihead))}{len(_irows)+1}"

ws5=wb.create_sheet("Prices Fort Collins")
ws5.append(["Ingredient","Key","Walmart $/100g","Costco $/100g","Cheaper at","kcal/100g","Protein/100g","$ per 25g protein"])
for c in ws5[1]:
    c.font=Font(name="Arial",bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="12365F")
    c.alignment=Alignment(vertical="center",wrap_text=True)
for k,(w_,c_) in sorted(PR.PRICE.items(), key=lambda kv: ING[kv[0]]["name"] if kv[0] in ING else kv[0]):
    if k not in ING: continue
    ig=ING[k]
    best = "Costco" if (c_ and c_ < w_) else "Walmart"
    per25 = round((25/ig["p"])*(w_/100)*100,2) if ig["p"] else ""
    ws5.append([ig["name"], k, w_, c_ if c_ else "", best, ig["kcal"], ig["p"], per25])
for r_ in ws5.iter_rows(min_row=2):
    for c in r_: c.font=Font(name="Arial")
for i,w_ in enumerate([36,26,16,16,12,12,14,18],1):
    ws5.column_dimensions[get_column_letter(i)].width=w_
ws5.freeze_panes="C2"
ws5.auto_filter.ref=f"A1:H{ws5.max_row}"

wb.save(f"{OUT}/meal_database.xlsx")

# ---------------- HTML ----------------
TAGCLASS = {
 "HIGH PROTEIN":"t-prot","VERY HIGH PROTEIN":"t-prot","HIGH LEUCINE":"t-leu","LEUCINE PRIORITY":"t-leu",
 "HIGH CALORIE":"t-cal","LOW CALORIE":"t-lowcal","HIGH FIBER":"t-fib","HIGH HEALTHY FAT":"t-fat",
 "HIGH CARB":"t-carb","LOW CARB":"t-carb","BALANCED MACRO":"t-bal","MUSCLE-BUILDING PRIORITY":"t-mus",
 "POST-WORKOUT FRIENDLY":"t-rec","PRE-WORKOUT FRIENDLY":"t-rec","HIGH SATIETY":"t-sat","QUICK":"t-quick",
 "TWO-MEAL-DAY FRIENDLY":"t-fast","BUDGET FRIENDLY":"t-budget","MEAL PREP":"t-budget",
 "FREEZER FRIENDLY":"t-budget","PORTABLE":"t-budget","VEGETARIAN":"t-veg","VEGAN":"t-veg",
 "SDA-COMPATIBLE":"t-veg","GLUTEN-FREE":"t-gf","OMEGA-3 RICH":"t-micro","HIGH MICRONUTRIENT DENSITY":"t-micro",
}
def tagcls(t):
    if t.startswith("HIGH ") and t.split()[-1] in ("CALCIUM","IRON","MAGNESIUM","POTASSIUM","ZINC"): return "t-micro"
    return TAGCLASS.get(t,"t-def")

def tagspan(t): return f'<span class="tag {tagcls(t)}">{html.escape(t)}</span>'

CATS = [("Breakfast","Breakfast Index","B"),("Lunch/Dinner","Lunch and Dinner Index","L"),
        ("Snack","Snack Index","S"),("Drink","Protein Drink Index","D"),
        ("SDA Meat/Fish","Optional SDA-Compatible Meat and Fish Section","M")]

def index_table(cat):
    out=['<table class="idx"><thead><tr><th>ID</th><th>Meal</th><th>Diff</th><th>Min</th><th>Cal</th>'
         '<th>P</th><th>C</th><th>F</th><th>Fib</th><th>Leu</th><th>Satiety</th><th>Key tags</th></tr></thead><tbody>']
    for d in DATA:
        if d["r"]["cat"]!=cat: continue
        p=d["per"]; keytags=[t for t in d["tags"] if t not in ("VEGETARIAN","SDA-COMPATIBLE","GLUTEN-FREE")][:5]
        out.append(f'<tr><td>{d["r"]["id"]}</td><td><a href="#{d["r"]["id"]}">{html.escape(d["r"]["name"])}</a></td>'
                   f'<td class="d-{d["diff"][:1]}">{d["diff"][:4].title()}</td><td>{d["time"]}</td>'
                   f'<td><b>{int(round(p["kcal"]/5.0)*5)}</b></td><td><b>{round(p["p"])}</b></td><td>{round(p["c"])}</td>'
                   f'<td>{round(p["f"])}</td><td>{round(p["fib"])}</td><td>{round(p["leu"],1)}</td>'
                   f'<td>{d["satlvl"].title()}</td><td class="tg">{" ".join(keytags)}</td></tr>')
    out.append("</tbody></table>")
    return "\n".join(out)

def recipe_block(d):
    r=d["r"]; p=d["per"]
    _cw,_cws,_ = PR.cost_of(r, r["servings"], "walmart")
    _cc,_ccs,_ = PR.cost_of(r, r["servings"], "costco")
    price=(f'<div class="pricetag">${_cws:.2f} per serving '
           f'<em>&middot; ${_cw:.2f} for all {r["servings"]} '
           f'&middot; <span data-store-label>Walmart</span></em></div>')
    ing="".join(f'<li><b>{g:g} g</b> {html.escape(ING[k]["name"])} <span class="hh">({html.escape(m)})</span></li>'
                for k,g,m in r["ing"])
    steps="".join(f"<li>{html.escape(s)}</li>" for s in r["steps"])
    subs="".join(f"<li>{html.escape(s)}</li>" for s in r.get("subs",[])) or "<li>None specific.</li>"
    var="".join(f"<li>{html.escape(s)}</li>" for s in r.get("variations",[])) or ""
    mi=", ".join(d["micros"]) or "Nothing above 15% DV."
    scale=[]
    for mult,lab in [(0.5,"0.5x"),(1,"1x"),(1.5,"1.5x"),(2,"2x")]:
        scale.append(f'<td>{lab}</td><td>{int(round(p["kcal"]*mult/5.0)*5)}</td><td>{round(p["p"]*mult)}</td>'
                     f'<td>{round(p["c"]*mult)}</td><td>{round(p["f"]*mult)}</td>')
    scrows="".join(f"<tr>{s}</tr>" for s in scale)
    return f"""
<div class="recipe" id="{r['id']}">
<h3><span class="rid">{r['id']}</span>{html.escape(r['name'])}</h3>
<p class="meta"><span class="d-{d['diff'][:1]}">{d['diff'].title()}</span> &middot;
Prep {r['prep']} min &middot; Cook {r['cook']} min &middot; Total {d['time']} min &middot;
Makes {r['servings']} serving{'s' if r['servings']>1 else ''} &middot; Satiety {d['satlvl'].title()} ({d['sat']})</p>
<p class="tags">{" ".join(tagspan(t) for t in d['tags'])}</p>
{price}<div class="macrobar">
<div><span>{int(round(p['kcal']/5.0)*5)}</span>kcal</div><div><span>{round(p['p'])} g</span>protein</div>
<div><span>{round(p['c'])} g</span>carbs</div><div><span>{round(p['f'])} g</span>fat</div>
<div><span>{round(p['fib'])} g</span>fiber</div><div><span>{round(p['leu'],1)} g</span>leucine</div>
</div>
<div class="cols">
<div><h4>Ingredients <span class="hh">(per {r['servings']} serving{'s' if r['servings']>1 else ''})</span></h4><ul class="ing">{ing}</ul></div>
<div><h4>Method</h4><ol class="steps">{steps}</ol></div>
</div>
<h4>Notable micronutrients</h4><p class="small">{html.escape(mi)}</p>
<h4>Storage</h4><p class="small">{html.escape(r.get('storage',''))}</p>
{f'<h4>Meal prep</h4><p class="small">{html.escape(r["prep_notes"])}</p>' if r.get('prep_notes') else ''}
<h4>Substitutions</h4><ul class="small">{subs}</ul>
{f'<h4>Variations</h4><ul class="small">{var}</ul>' if var else ''}
<h4>Scaling</h4>
<table class="scale"><tr><th>Portion</th><th>Cal</th><th>P</th><th>C</th><th>F</th></tr>{scrows}</table>
<p class="back"><a href="#idx-{r['cat'][:1]}">Back to index</a> &middot; <a href="#top">Top</a></p>
</div>"""

CSS = THEME_CSS + THEME_EXTRA + THEME_EXTRA2 + THEME_EXTRA3 + THEME_EXTRA4 + THEME_EXTRA5 + THEME_EXTRA6

def build_html(printable_only=False):
    parts=[f'<!DOCTYPE html><html><head><meta charset="utf-8"><title>The Meal Handbook</title>{CSS}</head><body><a id="top"></a>']
    parts.append(CALCU.COVER.format(n=len(DATA), ing=len(ING)))
    parts.append('<nav class="topnav">'
      '<a href="#howto">Start</a><a href="#live">Calculator</a><a href="#calc">Formulas</a>'
      '<a href="#targets">Your numbers</a><a href="#collections">Collections</a>'
      '<a href="#c-cheat">Cheat meals</a><a href="#c-dessert">Desserts</a>'
      '<a href="#c-fast">10 min</a><a href="#c-cheap">Cheap</a>'
      '<a href="#idx-B">Breakfast</a><a href="#idx-L">Mains</a><a href="#idx-S">Snacks</a>'
      '<a href="#idx-D">Drinks</a><a href="#idx-M">Meat</a>'
      '<a href="#grocery-rotation">Groceries</a><a href="#ingredients">Ingredients</a></nav>')
    counts = Counter(d["r"]["cat"] for d in DATA)
    parts.append('<h2 id="toc">Contents</h2><div class="toc">'
      '<a href="#howto">Start here</a>'
      '<a href="#assumptions">1. The ground rules</a>'
      '<a href="#household">Cooking for two</a>'
      '<a href="#app">The kitchen app</a>'
      '<a href="#recommend">Post-training picks</a>'
      '<a href="#lists">Favourites and lists</a>'
      '<a href="#live">Live calculator</a>'
      '<a href="#calc">The calculation engine</a>'
      '<a href="#targets">2. My numbers</a>'
      '<a href="#protein">3. Protein, EAAs, leucine</a>'
      '<a href="#cfw">4. Carbs, fat, fiber, water</a>'
      '<a href="#micros">5. Micronutrients</a>'
      '<a href="#timing">6. Timing, fasting, two-meal days</a>'
      '<a href="#goals">7. Goal guides</a>'
      '<a href="#tags">8. Tag definitions</a>'
      '<a href="#build">9. Build your own meal</a>'
      '<a href="#decide">10. What should I eat right now</a>'
      '<a href="#grocery">11. Groceries</a>'
      '<a href="#prep">12. Meal prep</a>'
      '<a href="#safety">13. Food safety</a>'
      '<a href="#relationship">14. Relationship with food</a>'
      '<a href="#tracking">15. Daily target page</a>'
      '<a href="#weekly">16. Weekly tracking</a>'
      '<a href="#supps">17. Supplements</a>'
      '<a href="#refs">18. References</a>'
      + "".join(f'<a href="#idx-{pre}">{title} ({counts[cat]})</a>' for cat,title,pre in CATS)
      + '<a href="#grocery-rotation">Four-week grocery rotation</a>'
      + '<a href="#collections">Collections: cheat meals, desserts, quick, cheap</a>'
      + '<a href="#ingredients">Master ingredient list</a>'
      + '<a href="#library">Full recipe library</a></div>')
    parts.append(f'''<div class="stripe">
<div><b>{len(DATA)}</b><span>Recipes</span></div>
<div><b>2,800</b><span>Target kcal / day</span></div>
<div><b>160 g</b><span>Target protein</span></div>
<div><b>365 g</b><span>Target carbs</span></div>
<div><b>78 g</b><span>Target fat</span></div>
<div><b>39 g</b><span>Target fiber</span></div>
<div><b>125 oz</b><span>Target water</span></div>
<div><b>+0.4 lb</b><span>Weekly goal</span></div>
</div>''')
    parts += [V.START_HERE, HH.HOUSEHOLD, APP.PANEL, CALCU.CALCULATOR, FRM.CALC, C.ASSUMPTIONS, C.TARGETS, C.PROTEIN, C.CARBS_FAT_FIBER, C.MICROS,
              C.TIMING, C.GOALS, C.SYSTEM, C.PRACTICAL, V.GROCERY_ROTATION, C.TRACKING]
    for cat,title,pre in CATS:
        parts.append(f'<h2 id="idx-{pre}">{title}</h2>')
        if pre=="M":
            parts.append('<p class="banner">This section is deliberately separate from the main catalog. '
             'Every item is SDA-compatible: no pork, no shellfish, no biblically unclean animals. '
             'Check processed products individually, since ingredients and processing aids vary by manufacturer. '
             'Nothing in the main catalog depends on anything in this section.</p>')
        parts.append(index_table(cat))
    parts.append(COLL.build_collections(DATA))
    _gh,_goff = COLL.gluten_audit(ING, DATA)
    parts.append(ILIST.build(ING, DATA, OUT)[0].replace('<h2 id="ingredients">Master ingredient list</h2>',
        '<h2 id="ingredients">Master ingredient list</h2>' + _gh))
    parts.append('<h2 id="library">Full recipe library</h2>')
    for cat,title,pre in CATS:
        n=sum(1 for d in DATA if d["r"]["cat"]==cat)
        parts.append(f'<div class="catrule">{title}<span>{n} recipes</span></div>')
        parts.append('<div class="rgrid">')
        for d in DATA:
            if d["r"]["cat"]==cat: parts.append(recipe_block(d))
        parts.append('</div>')
    _targets = json.dumps({
      "j":{"kcal":2800,"p":160,"c":365,"f":78,"fib":39,"w":125},
      "a":{"kcal":1900,"p":100,"c":250,"f":53,"fib":27,"w":85},
      "b":{"kcal":4700,"p":260,"c":615,"f":131,"fib":66,"w":210}})
    parts.append(APP.script(APP.data_blob(DATA, PR.cost_of), _targets))
    parts.append("</body></html>")
    return "\n".join(parts)

htm = build_html()
with open(f"{OUT}/nutrition_handbook.html","w") as fh: fh.write(htm)
print("HTML written")
